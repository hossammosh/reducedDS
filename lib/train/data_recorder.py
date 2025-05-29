import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Global variables to track initialization state and current epoch
_current_epoch = None
_epoch_files_initialized = {}  # Track which epoch files have been initialized


# Generate filename for a specific epoch
def _get_epoch_filename(epoch):
    return f'samples_log_epoch_{epoch}.xlsx'


# Initialize a new Excel workbook with header for a specific epoch
def init_excel_for_epoch(epoch):
    global _epoch_files_initialized, _current_epoch

    # Update the current epoch
    _current_epoch = epoch

    # Generate filename for this epoch
    filename = _get_epoch_filename(epoch)

    # Check if this epoch's file has already been initialized
    if epoch in _epoch_files_initialized and _epoch_files_initialized[epoch]:
        return

    # Create new workbook for this epoch
    wb = Workbook()
    ws = wb.active
    ws.title = "DataInfo"

    # Header row with specific headers
    headers = [
        "Index", "Sample Index", "stats/Loss_total", "stats_IoU", "Seq Name",
        "Template Frame ID", "Template Frame Path",
        "Search Frame ID",
        "Seq ID", "Seq Path", "Class Name", "Vid ID", "Search Names", "Search Path"
    ]
    ws.append(headers)
    _format_header_cells(ws)

    # Save the initial file
    wb.save(filename)
    _epoch_files_initialized[epoch] = True
    print(f"Excel file '{filename}' created with headers for epoch {epoch}.")


# Apply alignment and sizing to header row only
def _format_header_cells(ws):
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Format only the header row
    for cell in ws[1]:
        cell.alignment = align

    # Set column widths based on header content
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 4

    # Set row height for header
    ws.row_dimensions[1].height = 25


# Apply formatting to newly added rows
def _format_new_rows(ws, start_row, end_row):
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Format the new rows
    for row_num in range(start_row, end_row + 1):
        for cell in ws[row_num]:
            cell.alignment = align
        ws.row_dimensions[row_num].height = 25


# Get the next available row number
def _get_next_row(ws):
    return ws.max_row + 1


# Logging function - Now detects the epoch from LTRTrainer and saves to the appropriate file
def log_data(sample_index: int, data_info: dict, stats: dict):
    global _current_epoch

    # Get the current epoch from data_info if available (which comes from the LTRTrainer)
    if 'epoch' in data_info:
        epoch = data_info['epoch']
    else:
        # Use the global current epoch if not provided in data_info
        epoch = _current_epoch

    # If no epoch information is available, use epoch 0 as default
    if epoch is None:
        epoch = 0

    # Initialize file for this epoch if not already done
    init_excel_for_epoch(epoch)

    # Get filename for the current epoch
    filename = _get_epoch_filename(epoch)

    # Load existing workbook for this epoch
    wb = load_workbook(filename)
    ws = wb.active

    # Get the next available row
    next_row = _get_next_row(ws)
    start_row = next_row
    end_row = next_row + 1

    # Calculate the actual index (row number - 1 for header)
    data_index = next_row - 1

    # Merge columns vertically for the two new rows
    merge_columns = [1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14]
    for col in merge_columns:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    # Helper to safely convert lists to string (first element fallback)
    def safe_str_list(value, idx=0):
        if isinstance(value, list):
            if len(value) > idx:
                elem = value[idx]
                if isinstance(elem, list):
                    return ", ".join(map(str, elem))
                else:
                    return str(elem)
            else:
                return ""
        elif value is None:
            return ""
        else:
            return str(value)

    # Extract specific stats, handle missing keys
    loss_total = stats.get("Loss/total", None)
    iou = stats.get("IoU", None)

    # Write row 1 values (merged and unmerged)
    ws.cell(row=start_row, column=1, value=data_index)
    ws.cell(row=start_row, column=2, value=sample_index)
    ws.cell(row=start_row, column=3, value=loss_total)
    ws.cell(row=start_row, column=4, value=iou)
    ws.cell(row=start_row, column=5, value=data_info.get("seq_name", ""))
    ws.cell(row=start_row, column=6, value=safe_str_list(data_info.get("template_ids"), 0))
    ws.cell(row=start_row, column=7, value=safe_str_list(data_info.get("template_path"), 0))
    ws.cell(row=start_row, column=8, value=safe_str_list(data_info.get("search_id")))
    ws.cell(row=start_row, column=9, value=data_info.get("seq_id", ""))
    ws.cell(row=start_row, column=10, value=data_info.get("seq_path", ""))
    ws.cell(row=start_row, column=11, value=data_info.get("class_name", ""))
    ws.cell(row=start_row, column=12, value=data_info.get("vid_id", ""))
    ws.cell(row=start_row, column=13, value=", ".join(map(str, data_info.get("search_names", []))))
    ws.cell(row=start_row, column=14, value=", ".join(map(str, data_info.get("search_path", []))))

    # Write row 2 values (template frame IDs and paths)
    ws.cell(row=end_row, column=6, value=safe_str_list(data_info.get("template_ids"), 1))
    ws.cell(row=end_row, column=7, value=safe_str_list(data_info.get("template_path"), 1))

    # Format the newly added rows
    _format_new_rows(ws, start_row, end_row)

    # Save to file
    wb.save(filename)


# Used to explicitly set the current epoch (can be called at the beginning of each epoch)
def set_epoch(epoch_number):
    global _current_epoch
    _current_epoch = epoch_number

    # Initialize the excel file for this epoch if it doesn't exist yet
    init_excel_for_epoch(epoch_number)


# Optional: Function to reset all epoch logs
def reset_log():
    """Delete all existing epoch log files and reset initialization tracking."""
    global _current_epoch, _epoch_files_initialized

    # Reset globals
    _current_epoch = None

    # Delete any existing log files
    for epoch in _epoch_files_initialized.keys():
        filename = _get_epoch_filename(epoch)
        if os.path.exists(filename):
            os.remove(filename)

    # Reset initialization tracking
    _epoch_files_initialized = {}

    print("All epoch logs have been reset.")

# import os
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Alignment
# from openpyxl.utils import get_column_letter
#
# # Output Excel file name
# FILENAME = 'log_data.xlsx'
#
# # Global variable to track if file has been initialized
# _file_initialized = False
#
#
# # Initialize a new Excel workbook with header (called once at the beginning)
# def init_excel():
#     global _file_initialized
#
#     # Check if file already exists, if so, don't recreate it
#     if os.path.exists(FILENAME):
#         _file_initialized = True
#         return
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "DataInfo"
#
#     # Updated header row with specific headers 'stats/Loss/total' and 'stats/IoU'
#     headers = [
#         "Index", "Sample Index", "stats/Loss_total", "stats_IoU", "Seq Name",
#         "Template Frame ID", "Template Frame Path",
#         "Search Frame ID",
#         "Seq ID", "Seq Path", "Class Name", "Vid ID", "Search Names", "Search Path"
#     ]
#     ws.append(headers)
#     _format_header_cells(ws)
#
#     # Save the initial file
#     wb.save(FILENAME)
#     _file_initialized = True
#     #print(f"Excel file '{FILENAME}' created with headers.")
#
#
# # Apply alignment and sizing to header row only
# def _format_header_cells(ws):
#     align = Alignment(horizontal='center', vertical='center', wrap_text=True)
#
#     # Format only the header row
#     for cell in ws[1]:
#         cell.alignment = align
#
#     # Set column widths based on header content
#     for col in ws.columns:
#         max_length = 0
#         for cell in col:
#             try:
#                 if cell.value is not None:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         col_letter = get_column_letter(col[0].column)
#         ws.column_dimensions[col_letter].width = max_length + 4
#
#     # Set row height for header
#     ws.row_dimensions[1].height = 25
#
#
# # Apply formatting to newly added rows
# def _format_new_rows(ws, start_row, end_row):
#     align = Alignment(horizontal='center', vertical='center', wrap_text=True)
#
#     # Format the new rows
#     for row_num in range(start_row, end_row + 1):
#         for cell in ws[row_num]:
#             cell.alignment = align
#         ws.row_dimensions[row_num].height = 25
#
#
# # Get the next available row number
# def _get_next_row(ws):
#     return ws.max_row + 1
#
#
# # Logging function - Now appends data instead of overwriting
# def log_data(sample_index: int, data_info: dict, stats: dict):
#     global _file_initialized
#
#     # Initialize file if not already done
#     if not _file_initialized:
#         init_excel()
#
#     # Load existing workbook
#     wb = load_workbook(FILENAME)
#     ws = wb.active
#
#     # Get the next available row
#     next_row = _get_next_row(ws)
#     start_row = next_row
#     end_row = next_row + 1
#
#     # Calculate the actual index (row number - 1 for header)
#     data_index = next_row - 1
#
#     # Merge columns vertically for the two new rows
#     merge_columns = [1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14]
#     for col in merge_columns:
#         ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
#
#     # Helper to safely convert lists to string (first element fallback)
#     def safe_str_list(value, idx=0):
#         if isinstance(value, list):
#             if len(value) > idx:
#                 elem = value[idx]
#                 if isinstance(elem, list):
#                     return ", ".join(map(str, elem))
#                 else:
#                     return str(elem)
#             else:
#                 return ""
#         elif value is None:
#             return ""
#         else:
#             return str(value)
#
#     # Extract specific stats, handle missing keys
#     loss_total = stats.get("Loss/total", None)
#     iou = stats.get("IoU", None)
#
#     # Write row 1 values (merged and unmerged)
#     ws.cell(row=start_row, column=1, value=data_index)
#     ws.cell(row=start_row, column=2, value=sample_index)
#     ws.cell(row=start_row, column=3, value=loss_total)
#     ws.cell(row=start_row, column=4, value=iou)
#     ws.cell(row=start_row, column=5, value=data_info.get("seq_name", ""))
#     ws.cell(row=start_row, column=6, value=safe_str_list(data_info.get("template_ids"), 0))
#     ws.cell(row=start_row, column=7, value=safe_str_list(data_info.get("template_path"), 0))
#     ws.cell(row=start_row, column=8, value=safe_str_list(data_info.get("search_id")))
#     ws.cell(row=start_row, column=9, value=data_info.get("seq_id", ""))
#     ws.cell(row=start_row, column=10, value=data_info.get("seq_path", ""))
#     ws.cell(row=start_row, column=11, value=data_info.get("class_name", ""))
#     ws.cell(row=start_row, column=12, value=data_info.get("vid_id", ""))
#     ws.cell(row=start_row, column=13, value=", ".join(map(str, data_info.get("search_names", []))))
#     ws.cell(row=start_row, column=14, value=", ".join(map(str, data_info.get("search_path", []))))
#
#     # Write row 2 values (template frame IDs and paths)
#     ws.cell(row=end_row, column=6, value=safe_str_list(data_info.get("template_ids"), 1))
#     ws.cell(row=end_row, column=7, value=safe_str_list(data_info.get("template_path"), 1))
#
#     # Format the newly added rows
#     _format_new_rows(ws, start_row, end_row)
#
#     # Save to file
#     wb.save(FILENAME)
#     #print(f"Data appended to row {start_row}-{end_row} in '{FILENAME}'")
#
#
# # Optional: Function to manually initialize the Excel file at the start of training
# def initialize_training_log():
#     """Call this function at the beginning of your training to ensure the Excel file is created."""
#     init_excel()
#
#
# # Optional: Function to reset the log file (delete and recreate)
# def reset_log():
#     """Delete the existing log file and reinitialize."""
#     global _file_initialized
#     if os.path.exists(FILENAME):
#         os.remove(FILENAME)
#     _file_initialized = False
#     init_excel()
